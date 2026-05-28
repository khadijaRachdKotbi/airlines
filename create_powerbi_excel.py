#!/usr/bin/env python3
"""
Crée un fichier Excel complet pour Power BI avec toutes les données
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from pathlib import Path

DATA_DIR = "/home/esprit/airlLines_Project/powerbi_data"
OUTPUT_FILE = "/home/esprit/airlLines_Project/Airlines_PowerBI_Data.xlsx"

# Définition des couleurs
COLORS = {
    'header_bg': '4472C4',      # Bleu Power BI
    'header_text': 'FFFFFF',    # Blanc
    'kpi_good': '4CAF75',       # Vert
    'kpi_warning': 'FFC000',    # Jaune
    'kpi_bad': 'FF6B6B',        # Rouge
    'accent': 'E74C3C',         # Rouge accent
    'border': 'D0D0D0'          # Gris clair
}

def create_styled_header(ws, headers, row=1):
    """Crée un en-tête stylisé"""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color=COLORS['header_text'], size=11)
        cell.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin', color=COLORS['border']),
            right=Side(style='thin', color=COLORS['border']),
            top=Side(style='thin', color=COLORS['border']),
            bottom=Side(style='thin', color=COLORS['border'])
        )
        ws.column_dimensions[get_column_letter(col)].width = 20

def create_kpi_card_data(ws, kpis, start_row=2):
    """Crée des données pour les cartes KPI"""
    headers = ['KPI', 'Valeur', 'Objectif', 'Unité', 'Statut']
    create_styled_header(ws, headers, start_row)

    for i, kpi in enumerate(kpis, start=start_row + 1):
        ws.cell(row=i, column=1, value=kpi['name'])
        ws.cell(row=i, column=2, value=kpi['value'])
        ws.cell(row=i, column=3, value=kpi.get('target', '-'))
        ws.cell(row=i, column=4, value=kpi.get('unit', '-'))
        ws.cell(row=i, column=5, value=kpi['status'])

        # Couleur selon statut
        status_cell = ws.cell(row=i, column=5)
        if kpi['status'] == '✅':
            status_cell.fill = PatternFill(start_color=COLORS['kpi_good'], end_color=COLORS['kpi_good'], fill_type="solid")
        elif kpi['status'] == '🟡':
            status_cell.fill = PatternFill(start_color=COLORS['kpi_warning'], end_color=COLORS['kpi_warning'], fill_type="solid")
        elif kpi['status'] == '❌':
            status_cell.fill = PatternFill(start_color=COLORS['kpi_bad'], end_color=COLORS['kpi_bad'], fill_type="solid")

def load_csv_to_sheet(ws, csv_file, sheet_name=None):
    """Charge un fichier CSV dans une feuille"""
    csv_path = Path(DATA_DIR) / csv_file
    if not csv_path.exists():
        print(f"⚠️  Fichier introuvable: {csv_file}")
        return

    with open(csv_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # En-tête
    if lines:
        headers = lines[0].strip().split(',')
        create_styled_header(ws, headers)

        # Données
        for i, line in enumerate(lines[1:], start=2):
            values = line.strip().split(',')
            for j, val in enumerate(values, start=1):
                ws.cell(row=i, column=j, value=val)

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

def create_instructions_sheet(ws):
    """Crée la feuille d'instructions"""
    ws.title = "📋 Instructions"

    instructions = [
        ["Airlines Dashboard - Instructions Power BI"],
        [""],
        ["🚀 COMMENT CRÉER VOTRE DASHBOARD"],
        ["", "", ""],
        ["Étape 1: Ouvrir Power BI Desktop"],
        ["", "Téléchargez Power BI Desktop si nécessaire:"],
        ["", "https://www.microsoft.com/fr-fr/power-platform/products/power-bi-desktop"],
        [""],
        ["Étape 2: Importer les données"],
        ["", "1. Cliquez sur 'Obtenir des données'"],
        ["", "2. Sélectionnez 'Classeur Excel'"],
        ["", f"3. Naviguez vers: {Path(OUTPUT_FILE).name}"],
        ["", "4. Sélectionnez toutes les feuilles"],
        ["", "5. Cliquez sur 'Charger'"],
        [""],
        ["Étape 3: Créer les mesures DAX"],
        ["", "1. Cliquez sur 'Modélisation' > 'Nouvelle mesure'"],
        ["", "2. Copiez les formules depuis la feuille 'DAX_Measures'"],
        ["", "3. Collez dans la barre de formules"],
        [""],
        ["📊 PAGES DU DASHBOARD"],
        ["", "", ""],
        ["Page 1 - Executive"],
        ["", "• 4 Cartes KPI: Satisfaction par modèle"],
        ["", "• 1 Jauge: Score global vs objectif"],
        ["", "• 3 Cartes: Aéroports, Chatbot, ROI"],
        [""],
        ["Page 2 - Performance Modèles"],
        ["", "• Tableau comparatif des modèles"],
        ["", "• Graphique radar (Axes: Accuracy, Precision, Recall, F1, ROC-AUC)"],
        ["", "• Graphique barres: Feature importance"],
        [""],
        ["Page 3 - Analyse Sentiment"],
        ["", "• Utilisez les fichiers images wordcloud_*.png"],
        [""],
        ["Page 4 - Analyse Aéroports"],
        ["", "• Top 5 pires aéroports (satisfaction < 20%)"],
        ["", "• Top 5 meilleurs aéroports"],
        [""],
        ["Page 5 - Chatbot Insights"],
        ["", "• Graphique treemap: Volume par intention"],
        ["", "• Graphique barres: Taux de résolution"],
        [""],
        ["Page 6 - Business Insights"],
        ["", "• ROI projeté: +25%"],
        ["", "• Recommandations prioritaires"],
        [""],
        ["💡 CONSEILS"],
        ["", "• Utilisez les slicers pour filtrer par aéroport"],
        ["", "• Ajoutez des tooltips pour plus de détails"],
        ["", "• Utilisez des indicateurs de performance (KPIs)"],
        ["", "• Publiez sur Power BI Service pour partager"],
        [""],
        ["📁 FICHIERS SOURCES"],
        ["", "Toutes les données sont dans le dossier powerbi_data/"],
        [""],
        ["❓ SUPPORT"],
        ["", "Documentation Power BI:"],
        ["", "https://learn.microsoft.com/fr-fr/power-bi/"],
    ]

    # Titre principal
    ws['A1'].font = Font(bold=True, size=16, color='4472C4')
    ws['A1'].fill = PatternFill(start_color='E8F0F8', end_color='E8F0F8', fill_type="solid")
    ws.merge_cells('A1:C1')

    # Appliquer les instructions
    for i, row in enumerate(instructions, start=1):
        for j, cell in enumerate(row, start=1):
            if cell:
                ws.cell(row=i, column=j, value=cell)
                if j == 1 and cell and cell[0] in ['🚀', '📊', '💡', '📁', '❓']:
                    ws.cell(row=i, column=j).font = Font(bold=True, size=12)
                if 'Étape' in str(cell):
                    ws.cell(row=i, column=j).font = Font(bold=True, color='4472C4')

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 30

def create_dax_measures_sheet(ws):
    """Crée la feuille avec les mesures DAX"""
    ws.title = "📐 DAX_Measures"

    # En-tête
    headers = ['Mesure', 'Expression DAX', 'Description']
    create_styled_header(ws, headers)

    # Mesures DAX
    measures = [
        {
            'name': 'XGBoost_F1',
            'expression': '95.91',
            'description': 'Score F1 du modèle XGBoost'
        },
        {
            'name': 'SVM_F1',
            'expression': '91.21',
            'description': 'Score F1 du modèle SVM'
        },
        {
            'name': 'DistilBERT_F1',
            'expression': '89.19',
            'description': 'Score F1 du modèle DistilBERT'
        },
        {
            'name': 'Customer_Satisfaction_Target',
            'expression': '90',
            'description': 'Objectif de satisfaction'
        },
        {
            'name': 'Global_Satisfaction_Score',
            'expression': '([XGBoost_F1]*0.5)+([SVM_F1]*0.3)+([DistilBERT_F1]*0.2)',
            'description': 'Score global pondéré'
        },
        {
            'name': 'Critical_Airports_Count',
            'expression': 'CALCULATE(COUNTROWS(airport_satisfaction_filtered), [mean] < 0.20)',
            'description': "Nombre d'aéroports critiques"
        },
        {
            'name': 'Worst_Airport',
            'expression': 'MAXX(TOPN(1, airport_satisfaction_filtered, [mean], ASC), [origin_norm])',
            'description': "Aéroport le moins bien noté"
        },
        {
            'name': 'Chatbot_Total_Questions',
            'expression': 'SUM(chatbot_metrics[question_count])',
            'description': 'Total des questions chatbot'
        },
        {
            'name': 'Chatbot_Avg_Resolution',
            'expression': 'AVERAGE(chatbot_metrics[resolution_rate])',
            'description': 'Taux moyen de résolution chatbot'
        },
        {
            'name': 'ROI_3_Years_Value',
            'expression': '25',
            'description': 'ROI sur 3 ans en pourcentage'
        },
        {
            'name': 'Time_to_ROI_Value',
            'expression': '6',
            'description': 'Délai pour atteindre le ROI en mois'
        },
        {
            'name': 'Business_Status',
            'expression': 'IF([Global_Satisfaction_Score] >= [Customer_Satisfaction_Target], "✅ Atteint", IF([Global_Satisfaction_Score] >= 85, "🟡 En cours", "❌ Action requise"))',
            'description': 'Statut business global'
        }
    ]

    for i, measure in enumerate(measures, start=2):
        ws.cell(row=i, column=1, value=measure['name'])
        ws.cell(row=i, column=2, value=measure['expression'])
        ws.cell(row=i, column=3, value=measure['description'])

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 35

def create_kpis_sheet(ws):
    """Crée la feuille KPIs consolidée"""
    ws.title = "📈 KPIs"

    kpis = [
        {'name': 'Satisfaction XGBoost', 'value': '95.91%', 'target': '90%', 'unit': '%', 'status': '✅'},
        {'name': 'Satisfaction SVM', 'value': '91.21%', 'target': '90%', 'unit': '%', 'status': '✅'},
        {'name': 'Satisfaction DistilBERT', 'value': '89.19%', 'target': '85%', 'unit': '%', 'status': '✅'},
        {'name': 'Score Global', 'value': '93.50%', 'target': '90%', 'unit': '%', 'status': '✅'},
        {'name': 'Aéroports Critiques', 'value': '3', 'target': '0', 'unit': '', 'status': '❌'},
        {'name': 'Taux Résolution Chatbot', 'value': '85%', 'target': '90%', 'unit': '%', 'status': '🟡'},
        {'name': 'ROI 3 Ans', 'value': '+25%', 'target': '+25%', 'unit': '%', 'status': '🎯'},
        {'name': 'Délai Résultats', 'value': '6', 'target': '6', 'unit': 'mois', 'status': '🎯'},
        {'name': 'Pire Aéroport', 'value': 'Newcastle', 'target': '-', 'unit': '', 'status': '⚠️'},
        {'name': 'Meilleur Aéroport', 'value': 'KUL', 'target': '-', 'unit': '', 'status': '✅'},
        {'name': 'Top Feature', 'value': 'Online Boarding', 'target': '-', 'unit': '', 'status': '🔥'},
        {'name': 'Impact Top Feature', 'value': '31.3%', 'target': '-', 'unit': '%', 'status': '🔥'},
        {'name': 'Top Intention Chatbot', 'value': 'Reservation', 'target': '-', 'unit': '', 'status': '🤖'},
        {'name': 'Volume Chatbot', 'value': '115', 'target': '-', 'unit': 'questions', 'status': '📊'},
    ]

    create_kpi_card_data(ws, kpis)

def create_recommendations_sheet(ws):
    """Crée la feuille de recommandations"""
    ws.title = "💡 Recommandations"

    recommendations = [
        ["Priorité", "Action", "Impact", "Délai", "Responsable"],
        ["P0 - URGENT", "Améliorer Online Boarding", "31.3%", "1 mois", "UX Team"],
        ["P0 - URGENT", "Action sur Newcastle (satisfaction: 9.1%)", "Moyen", "2 mois", "Opérations"],
        ["P0 - URGENT", "Action sur Bogota (satisfaction: 18.2%)", "Moyen", "2 mois", "Opérations"],
        ["P0 - URGENT", "Action sur Karachi (satisfaction: 18.8%)", "Moyen", "2 mois", "Opérations"],
        ["P1 - HAUTE", "Optimiser Chatbot vers 90%", "Moyen", "3 mois", "AI Team"],
        ["P1 - HAUTE", "Continuer entraînement XGBoost", "Faible", "1 mois", "Data Science"],
        ["P2 - MOYENNE", "Améliorer Type of Travel Personal (impact: 22.9%)", "Moyen", "3 mois", "UX Team"],
        ["P2 - MOYENNE", "Améliorer In-flight Wifi (impact: 13.3%)", "Faible", "2 mois", "Tech Team"],
        ["P3 - FAIBLE", "Analyser les sentiments négatifs", "Faible", "Ongoing", "Analytics"],
        ["P3 - FAIBLE", "Développer des features pour les aéroports performants", "Faible", "6 mois", "Product"],
    ]

    # En-tête
    create_styled_header(ws, recommendations[0])

    # Données
    for i, rec in enumerate(recommendations[1:], start=2):
        for j, value in enumerate(rec, start=1):
            cell = ws.cell(row=i, column=j, value=value)

            # Couleur selon priorité
            if j == 1:
                if 'P0' in value:
                    cell.fill = PatternFill(start_color=COLORS['kpi_bad'], end_color=COLORS['kpi_bad'], fill_type="solid")
                elif 'P1' in value:
                    cell.fill = PatternFill(start_color=COLORS['kpi_warning'], end_color=COLORS['kpi_warning'], fill_type="solid")
                elif 'P2' in value:
                    cell.fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type="solid")

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20

def create_visual_config_sheet(ws):
    """Crée une feuille avec la configuration des visuels"""
    ws.title = "🎨 Visuels_Config"

    config = [
        ["Page", "Type Visuel", "Titre", "Source de données", "Champs"],
        ["Executive", "Card", "Satisfaction XGBoost", "Mesure DAX", "XGBoost_F1"],
        ["Executive", "Card", "Satisfaction SVM", "Mesure DAX", "SVM_F1"],
        ["Executive", "Card", "Satisfaction DistilBERT", "Mesure DAX", "DistilBERT_F1"],
        ["Executive", "Card", "Score Global", "Mesure DAX", "Global_Satisfaction_Score"],
        ["Executive", "Gauge", "Objectif Global", "Mesure DAX", "Global_Satisfaction_Score vs Customer_Satisfaction_Target"],
        ["Executive", "Card", "Aéroports Critiques", "Mesure DAX", "Critical_Airports_Count"],
        ["Executive", "Card", "Résolution Chatbot", "Mesure DAX", "Chatbot_Avg_Resolution"],
        ["Executive", "Card", "ROI 3 Ans", "Mesure DAX", "ROI_3_Years_Value"],
        ["Models", "Table", "Comparaison Modèles", "nlp_model_comparison", "Toutes colonnes"],
        ["Models", "Radar Chart", "Radar Comparatif", "nlp_model_comparison", "Model, Accuracy, Precision, Recall, F1-Score, ROC-AUC"],
        ["Models", "Bar Chart", "Feature Importance", "word_importance", "word (axe), score (valeur)"],
        ["Sentiment", "Image", "Mots Positifs", "wordcloud_positive.png", "Image externe"],
        ["Sentiment", "Image", "Mots Négatifs", "wordcloud_negative.png", "Image externe"],
        ["Airports", "Bar Chart", "Top 5 Pires", "airport_satisfaction_filtered", "origin_norm (axe), mean (valeur), filtre mean<0.25"],
        ["Airports", "Bar Chart", "Top 5 Meilleurs", "airport_satisfaction_filtered", "origin_norm (axe), mean (valeur), tri DESC"],
        ["Chatbot", "Treemap", "Volume par Intention", "chatbot_metrics", "intent (groupe), question_count (valeur)"],
        ["Chatbot", "Bar Chart", "Taux Résolution", "chatbot_metrics", "intent (axe), resolution_rate (valeur)"],
        ["Business", "Card", "ROI Projeté", "KPIs", "ROI 3 Ans"],
        ["Business", "Multi-row Card", "Recommandations", "Recommandations", "Priorité, Action, Impact"],
    ]

    create_styled_header(ws, config[0])

    for i, row in enumerate(config[1:], start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 50

def main():
    print("=" * 60)
    print("Création du fichier Excel pour Power BI")
    print("=" * 60)

    # Créer le classeur
    wb = openpyxl.Workbook()

    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Créer les feuilles de données (CSV)
    data_sheets = [
        ("XGBoost_Metrics", "xgboost_metrics.csv"),
        ("Models_Comparison", "nlp_model_comparison.csv"),
        ("Airports", "airport_satisfaction_filtered.csv"),
        ("Features", "word_importance.csv"),
        ("Chatbot", "chatbot_metrics.csv"),
        ("Business_KPIs", "business_kpis.csv"),
    ]

    for sheet_name, csv_file in data_sheets:
        print(f"📊 Création feuille: {sheet_name}")
        ws = wb.create_sheet(title=sheet_name)
        load_csv_to_sheet(ws, csv_file, sheet_name)

    # Créer les feuilles spéciales
    print(f"📋 Création feuille: Instructions")
    create_instructions_sheet(wb.create_sheet())

    print(f"📐 Création feuille: DAX Measures")
    create_dax_measures_sheet(wb.create_sheet())

    print(f"📈 Création feuille: KPIs")
    create_kpis_sheet(wb.create_sheet())

    print(f"💡 Création feuille: Recommandations")
    create_recommendations_sheet(wb.create_sheet())

    print(f"🎨 Création feuille: Configuration Visuels")
    create_visual_config_sheet(wb.create_sheet())

    # Sauvegarder
    wb.save(OUTPUT_FILE)
    print()
    print(f"✅ Fichier créé: {OUTPUT_FILE}")

    # Afficher le résumé
    print()
    print("=" * 60)
    print("RÉSUMÉ")
    print("=" * 60)
    print(f"Fichier: {Path(OUTPUT_FILE).name}")
    print(f"Taille: {Path(OUTPUT_FILE).stat().st_size / 1024:.1f} KB")
    print(f"Feuilles: {len(wb.sheetnames)}")
    print()
    print("📋 Feuilles créées:")
    for i, sheet in enumerate(wb.sheetnames, start=1):
        print(f"  {i}. {sheet}")
    print()
    print("🚀 PROCHAINES ÉTAPES:")
    print("  1. Ouvrez Power BI Desktop")
    print("  2. Obtenir des données > Classeur Excel")
    print(f"  3. Sélectionnez: {Path(OUTPUT_FILE).name}")
    print("  4. Sélectionnez toutes les feuilles")
    print("  5. Chargez et créez votre dashboard")
    print()

if __name__ == "__main__":
    main()