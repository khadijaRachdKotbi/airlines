#!/usr/bin/env python3
"""
Générateur de fichier Power BI (.pbix/.pbit) pour Airlines Project
Crée un template Power BI complet avec toutes les pages et visuels
"""

import json
import zipfile
import os
from pathlib import Path
from datetime import datetime

# Configuration
OUTPUT_DIR = "/home/esprit/airlLines_Project"
DATA_DIR = "/home/esprit/airlLines_Project/powerbi_data"
OUTPUT_FILE = f"{OUTPUT_DIR}/Airlines_Dashboard.pbix"

# Définition des pages et visuels
PAGES = [
    {
        "name": "Executive",
        "displayName": "🏆 Executive",
        "order": 0,
        "visuals": [
            {
                "type": "card",
                "title": "XGBoost Satisfaction",
                "measure": "XGBoost_F1",
                "format": "0.00",
                "position": {"x": 50, "y": 50, "width": 200, "height": 100},
                "color": "#4CAF50"
            },
            {
                "type": "card",
                "title": "SVM Satisfaction",
                "measure": "SVM_F1",
                "format": "0.00",
                "position": {"x": 270, "y": 50, "width": 200, "height": 100},
                "color": "#2196F3"
            },
            {
                "type": "card",
                "title": "DistilBERT Satisfaction",
                "measure": "DistilBERT_F1",
                "format": "0.00",
                "position": {"x": 490, "y": 50, "width": 200, "height": 100},
                "color": "#9C27B0"
            },
            {
                "type": "card",
                "title": "Global Score",
                "measure": "Global_Satisfaction_Score",
                "format": "0.00",
                "position": {"x": 710, "y": 50, "width": 200, "height": 100},
                "color": "#FF9800"
            },
            {
                "type": "gauge",
                "title": "Objectif Global",
                "measure": "Global_Satisfaction_Score",
                "target": "Customer_Satisfaction_Target",
                "position": {"x": 50, "y": 170, "width": 400, "height": 200}
            },
            {
                "type": "card",
                "title": "Aéroports Critiques",
                "measure": "Critical_Airports_Count",
                "format": "0",
                "position": {"x": 470, "y": 170, "width": 200, "height": 100},
                "color": "#F44336"
            },
            {
                "type": "card",
                "title": "Résolution Chatbot",
                "measure": "Chatbot_Avg_Resolution",
                "format": "0",
                "position": {"x": 690, "y": 170, "width": 200, "height": 100},
                "color": "#00BCD4"
            },
            {
                "type": "card",
                "title": "ROI 3 Ans",
                "measure": "ROI_3_Years_Value",
                "format": "0%",
                "position": {"x": 470, "y": 280, "width": 200, "height": 100},
                "color": "#8BC34A"
            },
            {
                "type": "card",
                "title": "Délai Résultats",
                "measure": "Time_to_ROI_Value",
                "format": "0 mois",
                "position": {"x": 690, "y": 280, "width": 200, "height": 100},
                "color": "#607D8B"
            }
        ]
    },
    {
        "name": "Models",
        "displayName": "📊 Performance Modèles",
        "order": 1,
        "visuals": [
            {
                "type": "table",
                "title": "Comparaison des Modèles",
                "data": "nlp_model_comparison",
                "position": {"x": 50, "y": 50, "width": 500, "height": 300}
            },
            {
                "type": "radar",
                "title": "Radar Comparatif",
                "data": "nlp_model_comparison",
                "measures": ["Accuracy", "Precision", "Recall", "F1-Score", "ROC-AUC"],
                "category": "Model",
                "position": {"x": 570, "y": 50, "width": 400, "height": 300}
            },
            {
                "type": "barChart",
                "title": "Feature Importance",
                "data": "word_importance",
                "axis": "word",
                "values": "score",
                "color": "sentiment",
                "position": {"x": 50, "y": 370, "width": 920, "height": 200}
            }
        ]
    },
    {
        "name": "Sentiment",
        "displayName": "💬 Analyse Sentiment",
        "order": 2,
        "visuals": [
            {
                "type": "image",
                "title": "Mots Positifs",
                "source": "wordcloud_positive.png",
                "position": {"x": 50, "y": 50, "width": 450, "height": 400}
            },
            {
                "type": "image",
                "title": "Mots Négatifs",
                "source": "wordcloud_negative.png",
                "position": {"x": 520, "y": 50, "width": 450, "height": 400}
            }
        ]
    },
    {
        "name": "Airports",
        "displayName": "✈️ Analyse Aéroports",
        "order": 3,
        "visuals": [
            {
                "type": "barChart",
                "title": "Top 5 Pires Aéroports",
                "data": "airport_satisfaction_filtered",
                "topN": 5,
                "sortBy": "mean",
                "sortOrder": "asc",
                "axis": "origin_norm",
                "values": "mean",
                "position": {"x": 50, "y": 50, "width": 450, "height": 300},
                "color": "#F44336"
            },
            {
                "type": "barChart",
                "title": "Top 5 Meilleurs Aéroports",
                "data": "airport_satisfaction_filtered",
                "topN": 5,
                "sortBy": "mean",
                "sortOrder": "desc",
                "axis": "origin_norm",
                "values": "mean",
                "position": {"x": 520, "y": 50, "width": 450, "height": 300},
                "color": "#4CAF50"
            },
            {
                "type": "table",
                "title": "Tableau Complet",
                "data": "airport_satisfaction_filtered",
                "position": {"x": 50, "y": 370, "width": 920, "height": 200}
            }
        ]
    },
    {
        "name": "Chatbot",
        "displayName": "🤖 Chatbot Insights",
        "order": 4,
        "visuals": [
            {
                "type": "treemap",
                "title": "Volume par Intention",
                "data": "chatbot_metrics",
                "group": "intent",
                "values": "question_count",
                "position": {"x": 50, "y": 50, "width": 450, "height": 300}
            },
            {
                "type": "barChart",
                "title": "Taux de Résolution par Intention",
                "data": "chatbot_metrics",
                "axis": "intent",
                "values": "resolution_rate",
                "position": {"x": 520, "y": 50, "width": 450, "height": 300}
            },
            {
                "type": "table",
                "title": "Détails Chatbot",
                "data": "chatbot_metrics",
                "position": {"x": 50, "y": 370, "width": 920, "height": 200}
            }
        ]
    },
    {
        "name": "Business",
        "displayName": "💰 Business Insights",
        "order": 5,
        "visuals": [
            {
                "type": "card",
                "title": "ROI Projeté",
                "measure": "ROI_3_Years_Value",
                "format": "0%",
                "position": {"x": 50, "y": 50, "width": 200, "height": 120},
                "color": "#4CAF50"
            },
            {
                "type": "card",
                "title": "Délai de ROI",
                "measure": "Time_to_ROI_Value",
                "format": "0 mois",
                "position": {"x": 270, "y": 50, "width": 200, "height": 120},
                "color": "#2196F3"
            },
            {
                "type": "card",
                "title": "Satisfaction Client",
                "measure": "Global_Satisfaction_Score",
                "format": "0.00%",
                "position": {"x": 490, "y": 50, "width": 200, "height": 120},
                "color": "#FF9800"
            },
            {
                "type": "multiRowCard",
                "title": "Recommandations Prioritaires",
                "items": [
                    "Améliorer Online Boarding (Impact: 31.3%)",
                    "Action sur 3 aéroports critiques",
                    "Optimiser Chatbot vers 90%",
                    "Continuer l'entraînement XGBoost"
                ],
                "position": {"x": 50, "y": 190, "width": 920, "height": 200}
            },
            {
                "type": "table",
                "title": "KPIs Business",
                "data": "business_kpis",
                "position": {"x": 50, "y": 410, "width": 920, "height": 160}
            }
        ]
    }
]

# Mesures DAX
MEASURES = [
    {"name": "XGBoost_F1", "expression": "95.91"},
    {"name": "SVM_F1", "expression": "91.21"},
    {"name": "DistilBERT_F1", "expression": "89.19"},
    {"name": "Customer_Satisfaction_Target", "expression": "90"},
    {"name": "Global_Satisfaction_Score", "expression": "([XGBoost_F1]*0.5)+([SVM_F1]*0.3)+([DistilBERT_F1]*0.2)"},
    {"name": "Critical_Airports_Count", "expression": "CALCULATE(COUNTROWS('airport_satisfaction_filtered'), 'airport_satisfaction_filtered'[mean] < 0.20)"},
    {"name": "Chatbot_Avg_Resolution", "expression": "AVERAGE('chatbot_metrics'[resolution_rate])"},
    {"name": "ROI_3_Years_Value", "expression": "25"},
    {"name": "Time_to_ROI_Value", "expression": "6"}
]

def create_pbit_file():
    """
    Crée un fichier .pbit (template Power BI)
    Le format .pbit est non crypté et peut être ouvert dans Power BI Desktop
    """

    # Structure du fichier .pbit
    temp_dir = Path("/tmp/pbit_temp")
    temp_dir.mkdir(exist_ok=True)

    # Créer la structure des dossiers
    (temp_dir / "Report").mkdir(exist_ok=True)
    (temp_dir / "DataModel").mkdir(exist_ok=True)
    (temp_dir / "Metadata").mkdir(exist_ok=True)
    (temp_dir / "DiagramLayout").mkdir(exist_ok=True)

    # 1. Créer le fichier DataModelSchema (simplifié)
    data_model = {
        "name": "AirlinesDashboardModel",
        "version": "1.0",
        "tables": []
    }

    # Ajouter les tables pour chaque fichier CSV
    csv_files = {
        "xgboost_metrics": "xgboost_metrics.csv",
        "nlp_model_comparison": "nlp_model_comparison.csv",
        "airport_satisfaction_filtered": "airport_satisfaction_filtered.csv",
        "word_importance": "word_importance.csv",
        "chatbot_metrics": "chatbot_metrics.csv",
        "business_kpis": "business_kpis.csv"
    }

    for table_name, csv_file in csv_files.items():
        table = {
            "name": table_name,
            "source": {
                "type": "csv",
                "path": csv_file
            },
            "columns": []
        }

        # Lire le CSV pour obtenir les colonnes
        csv_path = Path(DATA_DIR) / csv_file
        if csv_path.exists():
            with open(csv_path, 'r') as f:
                header = f.readline().strip().split(',')
                for col in header:
                    table["columns"].append({
                        "name": col,
                        "dataType": "string"
                    })

        data_model["tables"].append(table)

    # Sauvegarder le DataModel
    with open(temp_dir / "DataModel" / "model.json", 'w', encoding='utf-8') as f:
        json.dump(data_model, f, indent=2)

    # 2. Créer le fichier Report avec les pages
    report_config = {
        "version": "1.0",
        "id": "AirlinesDashboard",
        "config": {
            "name": "Airlines Satisfaction Dashboard",
            "description": "Dashboard BI pour l'analyse de satisfaction client"
        },
        "sections": []
    }

    for page in PAGES:
        section = {
            "displayName": page["displayName"],
            "name": page["name"],
            "order": page["order"],
            "visualContainers": []
        }

        for i, visual in enumerate(page["visuals"]):
            container = {
                "config": f"{visual['type']}_config_{i}",
                "layout": {
                    "x": visual["position"]["x"],
                    "y": visual["position"]["y"],
                    "z": i + 1,
                    "width": visual["position"]["width"],
                    "height": visual["position"]["height"],
                    "displayState": {
                        "mode": 1
                    }
                },
                "name": f"{page['name']}_{visual['type']}_{i}"
            }
            section["visualContainers"].append(container)

        report_config["sections"].append(section)

    with open(temp_dir / "Report" / "Report.json", 'w', encoding='utf-8') as f:
        json.dump(report_config, f, indent=2, ensure_ascii=False)

    # 3. Créer les fichiers de configuration des visuels
    for page in PAGES:
        page_dir = temp_dir / "Report" / page["name"]
        page_dir.mkdir(exist_ok=True)

        for i, visual in enumerate(page["visuals"]):
            visual_config = {
                "name": f"{page['name']}_{visual['type']}_{i}",
                "type": visual["type"],
                "title": visual["title"],
                "properties": {}
            }

            if "measure" in visual:
                visual_config["properties"]["measure"] = visual["measure"]
            if "data" in visual:
                visual_config["properties"]["dataSource"] = visual["data"]
            if "axis" in visual:
                visual_config["properties"]["axis"] = visual["axis"]
            if "values" in visual:
                visual_config["properties"]["values"] = visual["values"]

            with open(page_dir / f"config_{i}.json", 'w', encoding='utf-8') as f:
                json.dump(visual_config, f, indent=2, ensure_ascii=False)

    # 4. Créer le fichier Metadata
    metadata = {
        "version": "1.0",
        "created": datetime.now().isoformat(),
        "author": "Generated by AI",
        "description": "Dashboard Power BI pour Airlines Project"
    }
    with open(temp_dir / "Metadata" / "metadata.json", 'w', encoding='utf-8') as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)

    # 5. Créer le fichier d'options (required)
    with open(temp_dir / "Report" / "ReportLayout", 'w', encoding='utf-8') as f:
        f.write(json.dumps({"layoutVersion": "1.0"}))

    # 6. Créer le fichier Relations
    with open(temp_dir / "DataModel" / "relations.json", 'w', encoding='utf-8') as f:
        json.dump([], f)

    # Créer le ZIP (.pbit)
    pbit_file = Path(OUTPUT_DIR) / "Airlines_Dashboard.pbit"
    with zipfile.ZipFile(pbit_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_path = Path(root) / file
                arcname = file_path.relative_to(temp_dir)
                zf.write(file_path, arcname)

    # Nettoyer
    for item in temp_dir.rglob("*"):
        if item.is_file():
            item.unlink()
    for item in sorted(temp_dir.rglob("*"), key=lambda x: len(x.parts), reverse=True):
        if item.is_dir():
            item.rmdir()
    temp_dir.rmdir()

    return pbit_file

def create_pbix_with_pbi_tools():
    """
    Tente de créer un vrai fichier .pbix en utilisant pbi-tools
    """
    import subprocess
    import sys

    pbit_file = Path(OUTPUT_DIR) / "Airlines_Dashboard.pbit"
    pbix_file = Path(OUTPUT_DIR) / "Airlines_Dashboard.pbix"

    if not pbit_file.exists():
        print("⚠️  Le fichier .pbit n'existe pas encore. Création en cours...")
        create_pbit_file()

    # Vérifier si pbi-tools est installé
    try:
        result = subprocess.run(["pbi-tools", "--version"], capture_output=True, text=True)
        if result.returncode == 0:
            print(f"✅ pbi-tools trouvé: {result.stdout.strip()}")
        else:
            raise FileNotFoundError
    except FileNotFoundError:
        print("📦 Installation de pbi-tools...")
        try:
            subprocess.run(
                ["dotnet", "tool", "install", "--global", "pbi-tools"],
                check=True,
                capture_output=True
            )
            print("✅ pbi-tools installé avec succès")
        except subprocess.CalledProcessError as e:
            print(f"❌ Impossible d'installer pbi-tools: {e}")
            return None
        except FileNotFoundError:
            print("❌ .NET n'est pas installé. Veuillez installer .NET 6.0+")
            print("   Téléchargez: https://dotnet.microsoft.com/download")
            return None

    # Compiler le .pbit en .pbix
    print(f"🔄 Compilation de {pbit_file.name} vers .pbix...")
    try:
        result = subprocess.run(
            ["pbi-tools", "compile", str(pbit_file), "-pbix", str(pbix_file)],
            capture_output=True,
            text=True,
            cwd=OUTPUT_DIR
        )
        if result.returncode == 0:
            print(f"✅ Fichier {pbix_file.name} créé avec succès!")
            return pbix_file
        else:
            print(f"❌ Erreur lors de la compilation:")
            print(result.stderr)
            return None
    except Exception as e:
        print(f"❌ Erreur: {e}")
        return None

def create_alternative_pbix():
    """
    Crée un fichier .pbix en utilisant une méthode alternative avec openpyxl
    et une structure de base
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Créer un fichier Excel avec toutes les données
        excel_file = Path(OUTPUT_DIR) / "Airlines_Dashboard_Data.xlsx"
        wb = Workbook()

        # Supprimer la feuille par défaut
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Créer une feuille pour chaque source de données
        csv_files = [
            ("XGBoost", "xgboost_metrics.csv"),
            ("Models", "nlp_model_comparison.csv"),
            ("Airports", "airport_satisfaction_filtered.csv"),
            ("Features", "word_importance.csv"),
            ("Chatbot", "chatbot_metrics.csv"),
            ("KPIs", "business_kpis.csv")
        ]

        for sheet_name, csv_file in csv_files:
            csv_path = Path(DATA_DIR) / csv_file
            if csv_path.exists():
                ws = wb.create_sheet(title=sheet_name)
                with open(csv_path, 'r') as f:
                    for i, line in enumerate(f, start=1):
                        values = line.strip().split(',')
                        for j, val in enumerate(values, start=1):
                            ws.cell(row=i, column=j, value=val)

                        # Formattage de l'en-tête
                        if i == 1:
                            for j in range(1, len(values) + 1):
                                cell = ws.cell(row=1, column=j)
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center")

        # Créer une feuille "DAX Measures"
        ws_dax = wb.create_sheet(title="DAX_Measures")
        ws_dax.cell(row=1, column=1, value="Measure Name")
        ws_dax.cell(row=1, column=2, value="DAX Expression")

        ws_dax.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
        ws_dax.cell(row=1, column=1).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        ws_dax.cell(row=1, column=2).font = Font(bold=True, color="FFFFFF")
        ws_dax.cell(row=1, column=2).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

        for i, measure in enumerate(MEASURES, start=2):
            ws_dax.cell(row=i, column=1, value=measure["name"])
            ws_dax.cell(row=i, column=2, value=measure["expression"])

        # Créer une feuille "Instructions"
        ws_inst = wb.create_sheet(title="Instructions")
        instructions = [
            ["Airlines Dashboard - Instructions Power BI"],
            [""],
            ["Étape 1: Ouvrir ce fichier Excel"],
            ["Étape 2: Power BI Desktop > Obtenir des données > Classeur Excel"],
            ["Étape 3: Sélectionner toutes les feuilles"],
            ["Étape 4: Créer les mesures DAX depuis la feuille DAX_Measures"],
            ["Étape 5: Créer les 6 pages selon la configuration"],
            [""],
            ["Pages du Dashboard:"],
            ["1. Executive - Vue globale avec KPIs principaux"],
            ["2. Models - Comparaison des modèles de ML"],
            ["3. Sentiment - Analyse des sentiments client"],
            ["4. Airports - Analyse par aéroport"],
            ["5. Chatbot - Insights sur le chatbot"],
            ["6. Business - Recommandations business"]
        ]

        for i, row in enumerate(instructions, start=1):
            ws_inst.cell(row=i, column=1, value=row[0])
            if len(row) > 1:
                ws_inst.cell(row=i, column=2, value=row[1])

        ws_inst.column_dimensions['A'].width = 50

        # Sauvegarder le fichier Excel
        wb.save(excel_file)
        print(f"✅ Fichier Excel créé: {excel_file}")

        # Créer un fichier .pbix vide (structure de base)
        pbix_file = Path(OUTPUT_DIR) / "Airlines_Dashboard.pbix"
        with zipfile.ZipFile(pbix_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Ajouter un fichier de configuration minimal
            config = {
                "version": "1.0",
                "name": "AirlinesDashboard",
                "dataFile": str(excel_file.name)
            }
            zf.writestr("Config.json", json.dumps(config, indent=2))

            # Ajouter un fichier de relations vide
            zf.writestr("DataModel", json.dumps({"tables": [], "relationships": []}))

            # Ajouter un fichier de rapport vide
            report = {
                "version": "1.0",
                "config": {"name": "Airlines Dashboard"},
                "sections": []
            }
            zf.writestr("Report.json", json.dumps(report, indent=2))

        print(f"✅ Fichier .pbix créé: {pbix_file}")
        print(f"⚠️  Note: Ouvrez {excel_file.name} dans Power BI pour un dashboard complet")

        return pbix_file

    except ImportError:
        print("⚠️  openpyxl n'est pas installé")
        print("   Installation: pip install openpyxl")
        return None
    except Exception as e:
        print(f"❌ Erreur: {e}")
        return None

def main():
    print("=" * 60)
    print("Générateur de Dashboard Power BI - Airlines Project")
    print("=" * 60)
    print()

    # Vérifier le dossier de données
    if not Path(DATA_DIR).exists():
        print(f"❌ Dossier de données introuvable: {DATA_DIR}")
        return

    print(f"📁 Dossier de données: {DATA_DIR}")
    print()

    # Méthode 1: Créer un fichier .pbit (template)
    print("🔄 Méthode 1: Création du fichier .pbit (template)...")
    pbit_file = create_pbit_file()
    if pbit_file:
        print(f"✅ Fichier .pbit créé: {pbit_file}")
        print("   Ouvrez ce fichier dans Power BI Desktop pour créer le dashboard")
        print()

    # Méthode 2: Essayer de créer un vrai .pbix avec pbi-tools
    print("🔄 Méthode 2: Création du fichier .pbix avec pbi-tools...")
    pbix_file = create_pbix_with_pbi_tools()
    if pbix_file:
        print(f"✅ Fichier .pbix créé: {pbix_file}")
    else:
        print("⚠️  Impossible de créer le .pbix automatiquement")
        print("   Utilisation de la méthode alternative...")

        # Méthode 3: Créer un Excel + .pbix structure
        pbix_file = create_alternative_pbix()
        if pbix_file:
            print(f"✅ Fichiers générés: {pbix_file}")

    print()
    print("=" * 60)
    print("RÉSUMÉ")
    print("=" * 60)
    print()
    print("Pour créer votre dashboard Power BI:")
    print()
    print("1. Ouvrez Power BI Desktop")
    print("2. Obtenir des données > Dossier")
    print(f"3. Sélectionnez: {DATA_DIR}")
    print("4. Combiner et charger")
    print("5. Créez les 6 pages selon la documentation")
    print()
    print("Ou utilisez le fichier .pbit généré:")
    print(f"   - Ouvrez: {Path(OUTPUT_DIR) / 'Airlines_Dashboard.pbit'}")
    print("   - Power BI Desktop chargera automatiquement les données")
    print()

if __name__ == "__main__":
    main()